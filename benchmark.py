import torch
import time
import psutil
import math
import gzip
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from matplotlib import colors
import os
from datasets import load_dataset
from sklearn.metrics import accuracy_score
import gzip
import sys

# For accurate memory usage benchmarking
import gc

from bitnet import BitNetTransformer
from bitnet.at import AutoregressiveWrapper
from torch.utils.data import DataLoader, Dataset
import torch.nn.functional as F

MAX_BATCHES = 5000
SEQ_LEN = 2048
BATCH_SIZE = 8

benchmarks = [
    ('Perplexity (PPL)', lambda dataset_name: benchmark_ppl(model, device, dataset_name)),
    ('Latency (ms)', lambda dataset_name: benchmark_latency(model, device, dataset_name) * 1000),
    ('Memory (MB)', lambda dataset_name: benchmark_memory_usage(model, device, dataset_name)),
    ('Text Continuation Perplexity', lambda dataset_name: text_continuation_benchmark(model, device, dataset_name)),
    ('Next Character Prediction Accuracy', lambda dataset_name: next_char_prediction_benchmark(model, device, dataset_name)),
]

dataset_paths = {
    "enwik8": "./data/enwik8.gz",
    "enwik9": "./data/enwik9.gz",
    "wikitext-2": "wikitext-2-raw-v1",
    "wikitext-103": "wikitext-103-raw-v1"
}

benchmark_descriptions = {
    'Perplexity (PPL)': {
        'short': "Measures prediction quality; lower is better.",
        'detailed': "Calculated as exp(average cross-entropy loss). Measures how well the model predicts the next character in a sequence. Lower values indicate better predictions."
    },
    'Latency (ms)': {
        'short': "Average processing time per batch.",
        'detailed': "Measured by timing the forward pass of the model on a batch of data, averaged over multiple runs. Excludes data loading time."
    },
    'Memory (MB)': {
        'short': "Additional memory used during inference.",
        'detailed': "Calculated by measuring the difference in GPU memory usage before and after running the model. Includes model parameters and intermediate activations."
    },
    'Text Continuation Perplexity': {
        'short': "Evaluates text continuation quality.",
        'detailed': "Computed by having the model continue text from a given prefix and calculating the perplexity of the true continuation under the model's predictions."
    },
    'Text Continuation Character Accuracy': {
        'short': "Accuracy of character-level predictions in continuations.",
        'detailed': "Measures the proportion of characters correctly predicted by the model when continuing text from a given prefix."
    },
    'Text Continuation Character Similarity': {
        'short': "Average probability assigned to correct characters in continuations.",
        'detailed': "Calculates the mean probability that the model assigns to the correct character at each position when continuing text, even if it's not the top prediction."
    },
    'Bits Per Character (BPC)': {
        'short': "Compression efficiency; lower is better.",
        'detailed': "Calculated as log2(Perplexity) / 8. Represents the average number of bits needed to encode each character under the model's predictions."
    },
    'Compression Ratio': {
        'short': "Data compression capability; higher is better.",
        'detailed': "Computed as 8 / BPC. Indicates how much the model can theoretically compress the data compared to using 8 bits per character."
    },
    'Size (MB)': {
        'short': "Model file size on disk.",
        'detailed': "Measured by checking the file size of the saved model weights. Indicates the storage requirements for the model."
    },
    'Parameters': {
        'short': "Total learnable parameters in the model.",
        'detailed': "Counted by summing the number of elements in all parameter tensors of the model. Indicates model complexity and capacity."
    },
    'Training Dataset': {
        'short': "Dataset used to train the model.",
        'detailed': "Specifies the corpus used for training, which can affect the model's knowledge and biases."
    },
    'Next Character Prediction Accuracy': {
        'short': "Accuracy of next-character predictions.",
        'detailed': "Calculated by having the model predict the next character in a sequence and comparing it to the true next character. Averaged over many samples."
    },
    'Average Target Character Probability': {
        'short': "Average probability assigned to correct next character.",
        'detailed': "Computes the mean probability that the model assigns to the correct next character, even when it's not the top prediction. Provides a more nuanced view of model performance."
    }
}

def load_dataset_by_name(dataset_name):
    if dataset_name in ["enwik8", "enwik9"]:
        with gzip.open(dataset_paths[dataset_name], 'rb') as file:
            data = file.read()
        # Convert bytes to integers, capping at 255
        return torch.tensor([min(b, 255) for b in data], dtype=torch.long)
    else:
        dataset = load_dataset("wikitext", dataset_paths[dataset_name], split="test")
        text = " ".join(dataset["text"])
        # For word-level datasets, we'll tokenize later
        return text

def prepare_data_and_dataloader(dataset_name, seq_len=1024, batch_size=1):
    data = load_dataset_by_name(dataset_name)
    dataset = TextSamplerDataset(data, seq_len)
    dataloader = DataLoader(dataset, batch_size=batch_size, shuffle=False)
    return dataset, dataloader
    
def decode_token(token):
    return str(chr(max(32, token)))

def decode_tokens(tokens):
    return "".join(list(map(decode_token, tokens)))

class TextSamplerDataset(Dataset):
    def __init__(self, data, seq_len, is_byte_level=True):
        super().__init__()
        self.data = data
        self.seq_len = seq_len
        self.is_byte_level = is_byte_level
        
        if not is_byte_level:
            # Tokenize word-level data
            self.tokenizer = GPT2Tokenizer.from_pretrained('gpt2')
            self.encoded_data = torch.tensor(self.tokenizer.encode(self.data), dtype=torch.long)
        else:
            self.encoded_data = self.data

    def __getitem__(self, index):
        rand_start = torch.randint(0, self.encoded_data.size(0) - self.seq_len - 1, (1,))
        full_seq = self.encoded_data[rand_start : rand_start + self.seq_len + 1].long()
        return full_seq

    def __len__(self):
        return self.encoded_data.size(0) // self.seq_len

def check_existing_results(model_name, filename='bitnet_benchmark_results.xlsx'):
    try:
        wb = load_workbook(filename)
        if model_name in wb.sheetnames:
            sheet = wb[model_name]
            return {sheet.cell(row=i, column=1).value: sheet.cell(row=i, column=2).value 
                    for i in range(2, sheet.max_row + 1)}
    except FileNotFoundError:
        pass
    return {}

def calculate_model_parameters(model):
    # Load the state dict
    return sum(p.numel() for p in model.parameters())

def load_model(model_path):
    # Load the state dict
    state_dict = torch.load(model_path, map_location='cpu')
    
    # PRint state dict keys and shapes
    for key, value in state_dict.items():
        print(f"Key: {key}, Shape: {value.shape}")
    
    if model_path == "bitnet_final_model.pth":
        model = BitNetTransformer(num_tokens=256, dim=512, depth=8)
    elif model_path == "bitnet_300M_final_model.pth":
        model = BitNetTransformer(num_tokens=512, dim=1024, depth=18, heads=16, ff_mult=4)

    model = AutoregressiveWrapper(model, max_seq_len=1024)
    
    try:
        model.load_state_dict(state_dict)
    except RuntimeError as e:
        print(f"Error loading state dict: {e}")
        print("Model structure:")
        print(model)
    
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    try:
        model = model.to(device)
    except RuntimeError:
        print("CUDA out of memory. Falling back to CPU.")
        device = torch.device("cpu")
        model = model.to(device)
    
    print("Model structure:")
    print(model)
    
    return model, device

def benchmark_ppl(model, device, dataset_name):
    print(f"Benchmarking perplexity for {dataset_name}...")
    data = load_dataset_by_name(dataset_name)
    is_byte_level = dataset_name in ["enwik8", "enwik9"]
    dataset = TextSamplerDataset(data, seq_len=1024, is_byte_level=is_byte_level)
    dataloader = DataLoader(dataset, batch_size=1, shuffle=False)
    
    model.eval()
    total_loss = 0
    total_tokens = 0
    total_batches = len(dataloader)
    processed_batches = 0
    
    with torch.no_grad():
        print(f"Starting benchmark...")
        for batch_count, batch in enumerate(dataloader, 1):
            try:
                sys.stdout.write(f"\rBatch {batch_count}/{MAX_BATCHES}")
                sys.stdout.flush()
                
                batch = batch[0].to(device)
                input_ids = batch[:-1]
                target_ids = batch[1:]
                
                logits = model(input_ids.unsqueeze(0))
                
                if logits.dim() == 3:
                    logits = logits.squeeze(0)
                    
                    min_len = min(logits.size(0), target_ids.size(0))
                    logits = logits[:min_len]
                    target_ids = target_ids[:min_len]
                    
                    loss = F.cross_entropy(logits, target_ids, reduction='sum')
                else:
                    print(f"Unexpected logits dimension: {logits.dim()}. Skipping this batch.")
                    continue
                
                total_loss += loss.item()
                total_tokens += target_ids.numel()
                
                processed_batches += 1
                if processed_batches >= MAX_BATCHES:
                    break
                
            except RuntimeError as e:
                print(f"Error processing batch {batch_count}: {e}")
                continue
    
    if total_tokens == 0:
        print("No valid batches processed. Cannot compute perplexity.")
        return float('inf')
    
    perplexity = math.exp(total_loss / total_tokens)
    print(f"\nPerplexity for {dataset_name}: {perplexity:.4f}")
    return perplexity

def benchmark_latency(model, device, dataset_name):
    print(f"Benchmarking latency for {dataset_name}...")
    data = load_dataset_by_name(dataset_name)
    is_byte_level = dataset_name in ["enwik8", "enwik9"]
    dataset = TextSamplerDataset(data, seq_len=1024, is_byte_level=is_byte_level)
    dataloader = DataLoader(dataset, batch_size=1, shuffle=False)
    
    model.eval()
    latencies = []
    total_batches = len(dataloader)
    processed_batches = 0
    
    with torch.no_grad():
        print(f"Starting benchmark...")
        for batch_count, batch in enumerate(dataloader, 1):
            sys.stdout.write(f"\rBatch {batch_count}/{MAX_BATCHES}")
            sys.stdout.flush()
            
            batch = batch[0].to(device)
            input_ids = batch[:-1].unsqueeze(0)  # Add batch dimension and remove last token
            
            start_time = time.time()
            _ = model(input_ids)
            end_time = time.time()
            latencies.append(end_time - start_time)
            
            processed_batches += 1
            if processed_batches >= MAX_BATCHES:
                break
    
    avg_latency = sum(latencies) / len(latencies)
    print(f"\nAverage latency for {dataset_name}: {avg_latency * 1000:.2f} ms")
    return avg_latency

def benchmark_memory_usage(model, device, dataset_name):
    print(f"Benchmarking memory usage for {dataset_name}...")
    data = load_dataset_by_name(dataset_name)
    is_byte_level = dataset_name in ["enwik8", "enwik9"]
    dataset = TextSamplerDataset(data, seq_len=1024, is_byte_level=is_byte_level)
    dataloader = DataLoader(dataset, batch_size=1, shuffle=False)
    
    model.eval()
    gc.collect()
    torch.cuda.empty_cache()
    
    if device.type == 'cuda':
        torch.cuda.reset_peak_memory_stats(device)
        initial_memory = torch.cuda.memory_allocated(device)
    else:
        initial_memory = psutil.Process().memory_info().rss
    
    with torch.no_grad():
        for batch in dataloader:
            batch = batch[0].to(device)
            input_ids = batch[:-1].unsqueeze(0)  # Add batch dimension and remove last token
            _ = model(input_ids)
            break  # Only need to run one batch
    
    if device.type == 'cuda':
        final_memory = torch.cuda.max_memory_allocated(device)
    else:
        final_memory = psutil.Process().memory_info().rss
    
    memory_used = (final_memory - initial_memory) / (1024 * 1024)  # Convert to MB
    print(f"\nMemory usage for {dataset_name}: {memory_used:.2f} MB")
    return memory_used

def save_results_to_excel(results, filename='bitnet_benchmark_results.xlsx'):
    try:
        wb = load_workbook(filename)
    except FileNotFoundError:
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
    
    for result_key, model_results in results.items():
        sheet_name = f"{model_results['Model']} - {model_results['Benchmark Dataset']}"
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            sheet = wb.create_sheet(sheet_name)
        
        # Write headers if the sheet is empty
        if sheet.max_row == 1:
            headers = ["Metric", "Value", "Short Description", "Detailed Description"]
            for col, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col, value=header)
        
        # Write results
        row = sheet.max_row + 1
        for metric, value in model_results.items():
            if not metric.endswith("Description") and metric not in ['Model', 'Test Dataset']:
                sheet.cell(row=row, column=1, value=metric)
                sheet.cell(row=row, column=2, value=value)
                if f"{metric} Short Description" in model_results:
                    sheet.cell(row=row, column=3, value=model_results[f"{metric} Short Description"])
                if f"{metric} Detailed Description" in model_results:
                    sheet.cell(row=row, column=4, value=model_results[f"{metric} Detailed Description"])
                row += 1
        
        # Adjust column widths
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].auto_size = True
    
    wb.save(filename)
    print(f"Results saved to {filename}")
    
def get_model_size(model_path):
    return os.path.getsize(model_path) / (1024 * 1024)  # Convert bytes to MB

def save_results_to_image(results, filename='bitnet_benchmark_results.png'):
    # Define the table data
    models = list(results.keys())
    parameters = [results[model]['Parameters'] for model in models]
    size = [f"{results[model]['Size (MB)']:.2f}" for model in models]
    training_dataset = [f"{results[model]['Training Dataset']}" for model in models]
    memory = [f"{results[model]['Memory (MB)']:.2f}" for model in models]
    latency = [f"{results[model]['Latency (ms)']:.2f}" for model in models]
    ppl = [f"{results[model]['Perplexity (PPL)']:.2f}" for model in models]

    # Create the table data
    data = [models, parameters, training_dataset, size, memory, latency, ppl]
    column_labels = ['Models', 'Parameters', 'Training Dataset', 'Size (MB)↓', 'Memory (MB)↓', 'Latency (ms)↓', 'Perplexity (PPL)↓']

    # Create the figure and axis with transparent background
    fig, ax = plt.subplots() # 36x24 for larger poster
    fig.patch.set_alpha(0) # Make the background transparent

    # Hide axes
    ax.axis('off')

    # Create the table
    table = ax.table(cellText=list(zip(*data)),
                     colLabels=column_labels,
                     cellLoc='center',
                     loc='center')

    # Set font size and style
    table.auto_set_font_size(False)
    table.set_fontsize(24)

    # Add bold lines between columns and rows
    for key, cell in table.get_celld().items():
        cell.set_linewidth(3)
        cell.set_facecolor('white')
        cell.set_text_props(wrap=True)
        
    # Make column names larger and bold
    for (row, col), cell in table.get_celld().items():
        if row == 0:
            cell.set_text_props(weight='bold', fontsize=30)
    
    # Adjust column widths based on column labels
    table.auto_set_column_width(col=list(range(len(column_labels) + 1)))
    table.scale(1.5, 2)

    # Adjust layout and save the figure
    plt.tight_layout()
    plt.savefig(filename, dpi=1200, bbox_inches='tight', transparent=True)
    print(f"Results saved to {filename}")
    
import random

def text_continuation_benchmark(model, device, dataset_name, num_samples=100, prefix_length=64, continuation_length=64, seed=None):
    if seed is None:
        seed = random.randint(1, 10000)
    print(f"Benchmarking text continuation on {dataset_name} with seed {seed}...")
    
    data = load_dataset_by_name(dataset_name)
    
    random.seed(seed)
    torch.manual_seed(seed)
    
    model.eval()
    total_perplexity = 0
    total_char_accuracy = 0
    total_char_similarity = 0
    processed_samples = 0
    error_samples = 0
    
    vocab_size = model.net.emb.weight.shape[0]
    
    with torch.no_grad():
        while processed_samples < num_samples:
            try:
                start_idx = random.randint(0, len(data) - prefix_length - continuation_length - 1)
                prefix = data[start_idx:start_idx + prefix_length]
                true_continuation = data[start_idx + prefix_length:start_idx + prefix_length + continuation_length]
                
                input_ids = prefix.unsqueeze(0).to(device)
                true_continuation_ids = true_continuation.unsqueeze(0).to(device)
                
                input_ids = torch.clamp(input_ids, 0, vocab_size - 1)
                true_continuation_ids = torch.clamp(true_continuation_ids, 0, vocab_size - 1)
                
                output = model.generate(input_ids, continuation_length)
                generated_continuation = output[0, prefix_length:]
                
                output_for_true = model(torch.cat([input_ids, true_continuation_ids], dim=1))
                min_len = min(output_for_true.size(1), true_continuation_ids.size(1) + prefix_length)
                output_for_true = output_for_true[:, prefix_length:min_len, :]
                true_continuation_ids = true_continuation_ids[:, :min_len-prefix_length]
                
                loss = F.cross_entropy(output_for_true.view(-1, output_for_true.size(-1)), true_continuation_ids.view(-1))
                perplexity = torch.exp(loss).item()
                
                probabilities = F.softmax(output_for_true, dim=-1)
                predicted_ids = output_for_true.argmax(dim=-1)
                char_accuracy = (predicted_ids == true_continuation_ids).float().mean().item()
                char_similarity = probabilities.gather(-1, true_continuation_ids.unsqueeze(-1)).mean().item()
                
                total_perplexity += perplexity
                total_char_accuracy += char_accuracy
                total_char_similarity += char_similarity
                processed_samples += 1
                
                if processed_samples <= 3 or processed_samples % 10 == 0:
                    print(f"\nSample {processed_samples}:")
                    print(f"Prefix: {''.join([chr(i) for i in prefix.tolist()])}")
                    print(f"True continuation: {''.join([chr(i) for i in true_continuation.tolist()])}")
                    print(f"Generated continuation: {''.join([chr(i) for i in generated_continuation.tolist()])}")
                    print(f"Perplexity: {perplexity:.4f}")
                    print(f"Character-level accuracy: {char_accuracy:.4f}")
                    print(f"Character-level similarity: {char_similarity:.4f}")
                    print(f"Input shape: {input_ids.shape}")
                    print(f"True continuation shape: {true_continuation_ids.shape}")
                    print(f"Model output shape: {output_for_true.shape}")
            
            except Exception as e:
                print(f"\nError processing sample: {str(e)}")
                error_samples += 1
                continue
    
    if processed_samples == 0:
        print("No samples were successfully processed.")
        return float('inf'), 0.0, 0.0
    
    avg_perplexity = total_perplexity / processed_samples
    avg_char_accuracy = total_char_accuracy / processed_samples
    avg_char_similarity = total_char_similarity / processed_samples
    
    print(f"\nAverage perplexity on text continuation: {avg_perplexity:.4f}")
    print(f"Average character-level accuracy: {avg_char_accuracy:.4f}")
    print(f"Average character-level similarity: {avg_char_similarity:.4f}")
    print(f"Processed {processed_samples} out of {num_samples} samples")
    print(f"Encountered errors in {error_samples} samples")
    
    return avg_perplexity, avg_char_accuracy, avg_char_similarity

def next_char_prediction_benchmark(model, device, dataset_name, num_samples=1000, sequence_length=128, seed=42):
    print(f"Benchmarking next-character prediction accuracy on {dataset_name} with seed {seed}...")
    
    data = load_dataset_by_name(dataset_name)
    
    random.seed(seed)
    torch.manual_seed(seed)
    
    model.eval()
    total_correct = 0
    total_chars = 0
    total_similarity = 0
    processed_samples = 0
    error_samples = 0
    
    vocab_size = model.net.emb.weight.shape[0]
    
    with torch.no_grad():
        while processed_samples < num_samples:
            try:
                start_idx = random.randint(0, len(data) - sequence_length - 1)
                input_seq = data[start_idx:start_idx + sequence_length]
                target_char = data[start_idx + sequence_length].item()
                
                input_ids = input_seq.unsqueeze(0).to(device)
                target_id = torch.tensor([target_char], dtype=torch.long).to(device)
                
                input_ids = torch.clamp(input_ids, 0, vocab_size - 1)
                target_id = torch.clamp(target_id, 0, vocab_size - 1)
                
                output = model(input_ids)
                logits = output[0, -1, :]
                probabilities = F.softmax(logits, dim=0)
                
                predicted_id = logits.argmax().item()
                target_probability = probabilities[target_id].item()
                
                is_correct = predicted_id == target_id.item()
                if is_correct:
                    total_correct += 1
                total_chars += 1
                total_similarity += target_probability
                processed_samples += 1
                
                if processed_samples <= 3 or processed_samples % 100 == 0:
                    print(f"\nSample {processed_samples}:")
                    print(f"Input sequence: {''.join([chr(i) for i in input_seq.tolist()])}")
                    print(f"Target character: '{chr(target_char)}' (ASCII: {target_char})")
                    print(f"Predicted character: '{chr(predicted_id)}' (ASCII: {predicted_id})")
                    print(f"Correct: {is_correct}")
                    print(f"Target probability: {target_probability:.4f}")
                    print(f"Input shape: {input_ids.shape}")
                    print(f"Output shape: {output.shape}")
            
            except Exception as e:
                print(f"\nError processing sample: {str(e)}")
                error_samples += 1
                continue
    
    if total_chars == 0:
        print("No characters were successfully processed.")
        return 0.0, 0.0
    
    accuracy = total_correct / total_chars
    avg_similarity = total_similarity / total_chars
    print(f"\nNext-character prediction accuracy: {accuracy:.4f}")
    print(f"Average target character probability: {avg_similarity:.4f}")
    print(f"Total characters predicted: {total_chars}")
    print(f"Processed {processed_samples} out of {num_samples} samples")
    print(f"Encountered errors in {error_samples} samples")
    
    return accuracy, avg_similarity
    
if __name__ == "__main__":
    # model_paths = ["bitnet_final_model.pth"]
    model_paths = ["bitnet_300M_final_model.pth", "bitnet_final_model.pth"]
    datasets = dataset_paths.keys()
    
    sys.stdout.flush()
    sys.stderr.flush()
    
    results = {}
    for model_path in model_paths:
        for dataset_name in datasets:
            train_dataset = ""
            parameters = ""
            model_name = ""
            
            if model_path == "bitnet_final_model.pth":
                model_name = "SU-BitNet b1.58 Small"
                train_dataset = "enwik8"
                parameters = "25M"
                
            elif model_path == "bitnet_300M_final_model.pth":
                model_name = "SU-BitNet b1.58 Large"
                train_dataset = "enwik9"
                parameters = "230M"
            
            print(f"\n\nBenchmarking model: {model_name} on {dataset_name}\n\n")
            
            result_key = f"{model_name} - {dataset_name}"
            
            # Prepare data 
            # with gzip.open(f"./data/{dataset_name}.gz") as file:
            #     X = np.fromstring(file.read(int(95e6)), dtype=np.uint8)
            #     _, vaX = np.split(X, [int(90e6)])
            #     data_val = torch.from_numpy(vaX)
            
            # Create dataset and dataloader
            SEQ_LEN = 1024
            print(f"Loading model...")
            model, device = load_model(model_path)
            print(f"Model loaded successfully on device: {device}.")

            dataset, dataloader = prepare_data_and_dataloader(dataset_name)

            print(f"Dataset: {dataset}")
            print(f"Initial Dataloader: {dataloader}")

            # Move data to the same device as the model
            dataloader = DataLoader(dataset, batch_size=1, shuffle=False, 
                                    collate_fn=lambda x: [i.to(device) for i in x])

            print(f"Device-aware Dataloader: {dataloader}")
                        
            # existing_results = check_existing_results(model_name)
            # results[model_name] = existing_results.copy()
            results[result_key] = {}
            results[result_key]['Model'] = model_name
            results[result_key]['Benchmark Dataset'] = dataset_name
            results[result_key]['Parameters'] = f"{calculate_model_parameters(model) / 1e6:.2f}M"
            results[result_key]['Training Dataset'] = train_dataset
            results[result_key]['Size (MB)'] = get_model_size(model_path)
            
            try:
                accuracy, avg_prob = next_char_prediction_benchmark(model, device, dataset_name)
                results[result_key]['Next Character Prediction Accuracy'] = accuracy
                results[result_key]['Average Target Character Probability'] = avg_prob
            except RuntimeError as e:
                print(f"Error benchmarking next character prediction: {e}")
                results[result_key]['Next Character Prediction Accuracy'] = float('inf')
                results[result_key]['Average Target Character Probability'] = float('inf')
                
            try:
                text_cont_ppl, char_acc, char_sim = text_continuation_benchmark(model, device, dataset_name)
                results[result_key]['Text Continuation Perplexity'] = text_cont_ppl
                results[result_key]['Text Continuation Character Accuracy'] = char_acc
                results[result_key]['Text Continuation Character Similarity'] = char_sim
            except RuntimeError as e:
                print(f"Error benchmarking text continuation: {e}")
                results[result_key]['Text Continuation Perplexity'] = float('inf')
                results[result_key]['Text Continuation Character Accuracy'] = float('inf')
                results[result_key]['Text Continuation Character Similarity'] = float('inf')
                
            
            for metric, benchmark_func in benchmarks:
                if metric not in results[result_key]:
                    print(f"Benchmarking {metric}...")
                    try:
                        results[result_key][metric] = benchmark_func(dataset_name)
                    except RuntimeError as e:
                        print(f"Error benchmarking {metric}: {e}")
                        results[result_key][metric] = float('inf')
                else:
                    print(f"Using existing result for {metric}: {results[result_key][metric]}")
                    
            # Special case for BPC as it's derived from PPL
            if 'Perplexity (PPL)' in results[result_key] and 'Bits Per Character (BPC)' not in results[result_key]:
                results[result_key]['Bits Per Character (BPC)'] = math.log2(results[result_key]['Perplexity (PPL)']) / 8
            
            if 'Bits Per Character (BPC)' in results[result_key] and 'Compression Ratio' not in results[result_key]:
                results[result_key]['Compression Ratio'] = 8 / results[result_key]['Bits Per Character (BPC)']
            
            print(f"Benchmarking complete for {model_name} on {dataset_name}")
            time.sleep(2)
            print("Continuing to next model/dataset...")
            time.sleep(1)
            
            # Save results after each model to prevent data loss if the script is interrupted
            save_results_to_excel({result_key: results[result_key]})
    
    for result_key in results:
        for metric in results[result_key]:
            if metric in benchmark_descriptions:
                results[result_key][f"{metric} Short Description"] = benchmark_descriptions[metric]['short']
                results[result_key][f"{metric} Detailed Description"] = benchmark_descriptions[metric]['detailed']
    
    # Final save of all results
    save_results_to_excel(results)
    # save_results_to_image(results)
    
    for result_key in results:
        print(f"=====================")
        print(f"Model: {results[result_key]['Model']}")
        print(f"Test Dataset: {results[result_key]['Test Dataset']}")
        print(f"Training Dataset: {results[result_key]['Training Dataset']}")
        for metric, value in results[result_key].items():
            if not metric.endswith("Description") and metric not in ['Model', 'Test Dataset', 'Training Dataset']:
                print(f"{metric}: {value}")
                if f"{metric} Short Description" in results[result_key]:
                    print(f"  Short: {results[result_key][f'{metric} Short Description']}")
                if f"{metric} Detailed Description" in results[result_key]:
                    print(f"  Detailed: {results[result_key][f'{metric} Detailed Description']}")
        print(f"=====================")
        print()